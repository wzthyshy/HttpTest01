from openpyxl import Workbook, load_workbook


class PracticeExcel:
    def create_data(self):
        """
        创建excel数据
        :return:
        """
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "create"
        ws1["A1"] = "身高"
        ws1["B1"] = "体重"
        height = [180, 160, 170, 155]
        weight = [60, 55, 60, 80]
        data_dict = {180:61, 160:51 , 170:61, 155:81}
        data_keys = [i for i in data_dict.keys()]

        for i in range(len(height)):
            ws1.cell(row=i+2, column=1).value = data_keys[i]
            ws1.cell(row=i+2, column=2).value = data_dict[data_keys[i]]
        # #读取列表
        # for i in range(len(height)):
        #     ws1.cell()
        #     ws1.cell(row=i+2, column=1).value = height[i]
        #     ws1.cell(row=i+2, column=2).value = weight[i]
        wb.save("data.xlsx")

    def get_data(self):
        ld = load_workbook(filename="data.xlsx")
        sheet = ld["create"]
        for i in range(5):
            print(sheet.cell(row=i + 1, column=1).value)

    def health_data(self):
        ld = load_workbook(filename="data.xlsx")
        sheet = ld["create"]
        sheet["C1"] = "备注"

        for i in range(4):
            height = sheet.cell(row=i + 2, column=1).value
            weight = sheet.cell(row=i + 2, column=2).value

            health_weight = (height - 70) * 0.6
            if weight == health_weight:
                print("这是健康的体重",weight)
                sheet.cell(row=i + 2, column=3).value = "健康体重"
        ld.save(filename="data.xlsx")


pe = PracticeExcel()
pe.health_data()